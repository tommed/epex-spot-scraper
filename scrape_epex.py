# path: scripts/scrape_epex.py
#!/usr/bin/env python3
"""
Scrape the EPEX Spot GB 30-min Continuous market results table
and write data rows into the first sheet of a provided XLSX template.

- Extracts numeric data from `.js-table-values table tbody tr`
- Ignores rows containing only hyphens (hour headers)
- Assumes ascending half-hour rows (HH 1..48)
- Writes from cell B1 onwards (row A is left for headers)

Presumed Python 3.12.8
"""

import argparse
import logging
from dataclasses import dataclass
from typing import List, Optional, Sequence

from openpyxl import load_workbook
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeoutError

PROD_HH = 30
MARKET_GB = "GB"
EPEX_MARKET = MARKET_GB
EPEX_PRODUCT = PROD_HH
CSS_TABLE_SELECTOR = "table.table-01 tbody tr"


@dataclass(frozen=True)
class HHRow:
    hh: int
    low: Optional[float]
    high: Optional[float]
    last: Optional[float]
    weight_avg: Optional[float]
    buy_volume: Optional[float]
    sell_volume: Optional[float]
    volume: Optional[float]


def as_float_or_none(s: str) -> Optional[float]:
    """Convert text to float or return None for '-' / empty."""
    s = (s or "").strip().replace(",", "")
    if not s or s == "-":
        return None
    try:
        return float(s)
    except ValueError:
        return None


def extract_rows_from_dom(page) -> List[HHRow]:
    """Extract numeric rows from the rendered EPEX table."""
    js = r"""
    (selector) => {
      const rows = [];
      const tableRows = document.querySelectorAll(selector);
      for (const tr of tableRows) {
        const tds = Array.from(tr.querySelectorAll('td')).map(td => (td.textContent || '').trim());
        // Skip if no data or all are '-'
        if (tds.length < 7) continue;
        const meaningful = tds.filter(v => v && v !== '-');
        if (meaningful.length === 0) continue;
        rows.push(tds.slice(0, 7));  // only first 7 numeric cols
      }
      return rows;
    }
    """
    raw_rows = page.evaluate(js, CSS_TABLE_SELECTOR)
    hh_rows: List[HHRow] = []
    for i, cells in enumerate(raw_rows, start=1):
        nums = [as_float_or_none(x) for x in cells]
        while len(nums) < 7:
            nums.append(None)
        hh_rows.append(
            HHRow(
                hh=i,
                low=nums[0],
                high=nums[1],
                last=nums[2],
                weight_avg=nums[3],
                buy_volume=nums[4],
                sell_volume=nums[5],
                volume=nums[6],
            )
        )
    logging.info("Extracted %d valid rows", len(hh_rows))
    return hh_rows


def write_rows_to_template(
    template_path: str, out_path: str, rows: Sequence[HHRow], start_row: int = 2, start_col: int = 1
) -> None:
    """Write rows into the provided XLSX template."""
    wb = load_workbook(template_path)
    ws = wb.worksheets[0]

    for i, r in enumerate(rows, start=start_row):
        ws.cell(row=i, column=start_col + 0, value=r.hh)
        ws.cell(row=i, column=start_col + 1, value=r.low)
        ws.cell(row=i, column=start_col + 2, value=r.high)
        ws.cell(row=i, column=start_col + 3, value=r.last)
        ws.cell(row=i, column=start_col + 4, value=r.weight_avg)
        ws.cell(row=i, column=start_col + 5, value=r.buy_volume)
        ws.cell(row=i, column=start_col + 6, value=r.sell_volume)
        ws.cell(row=i, column=start_col + 7, value=r.volume)

    wb.save(out_path)


def run(url: str, template: str, out: str, timeout_ms: int = 30000) -> None:
    logging.info("Navigating to %s", url)
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True, args=["--disable-blink-features=AutomationControlled"])
        context = browser.new_context(user_agent=(
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/120.0.0.0 Safari/537.36"
        ))
        page = context.new_page()

        # Debugging file
        #html = page.content()
        #with open("debug.html", "w", encoding="utf-8") as f:
        #    f.write(html)

        # Scrape procedure
        try:
            page.goto(url, timeout=timeout_ms)
            # Wait for the table to be present (not necessarily visible)
            page.wait_for_selector(CSS_TABLE_SELECTOR, state="attached", timeout=timeout_ms)

            # Attempt to close cookie banner if present
            try:
                consent_button = page.query_selector("button:has-text('Accept')") or page.query_selector("button:has-text('Agree')")
                if consent_button:
                    consent_button.click()
                    logging.info("Accepted cookie banner.")
            except Exception:
                pass

            # Give the page a short grace period for CSS display changes
            page.wait_for_timeout(1000)
        except PlaywrightTimeoutError as e:
            logging.error("Timed out loading table node: %s", e)
            html_preview = page.content()[:2000]
            logging.debug("Page HTML snippet:\n%s", html_preview)
            raise SystemExit(2)

        # Extract
        rows = extract_rows_from_dom(page)
        if not rows:
            logging.error("No data rows found; verify table selector or structure.")
            raise SystemExit(3)

        # Load
        write_rows_to_template(template, out, rows)
        logging.info("Wrote %d rows to %s", len(rows), out)

        # Teardown
        context.close()
        browser.close()



def parse_args():
    p = argparse.ArgumentParser(description="Scrape EPEX Spot GB 30-min table into an XLSX template")
    p.add_argument("--date", required=True, help="The date to grab data for")
    p.add_argument("--template", required=True, help="Path to input XLSX template")
    p.add_argument("--out", required=True, help="Output XLSX path")
    p.add_argument("--timeout-ms", type=int, default=30000, help="Page load timeout (ms)")
    p.add_argument("--log-level", default="INFO", help="Logging level")
    return p.parse_args()

def get_epex_url(market, date, product):
    return f"https://www.epexspot.com/en/market-results?market_area={EPEX_MARKET}&delivery_date={args.date}&modality=Continuous&data_mode=table&product={EPEX_PRODUCT}"


if __name__ == "__main__":
    args = parse_args()
    logging.basicConfig(level=getattr(logging, args.log_level.upper(), logging.INFO),
                        format="%(levelname)s: %(message)s")
    url = get_epex_url(EPEX_MARKET, args.date, EPEX_PRODUCT)
    run(url, args.template, args.out, timeout_ms=args.timeout_ms)
