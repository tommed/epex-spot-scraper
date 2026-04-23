# path: scripts/scrape_epex_da.py
#!/usr/bin/env python3
"""
Scrape the EPEX Spot GB Day Ahead Auction market results table
and write data rows into the first sheet of a provided XLSX template.

- Extracts numeric data from `table.table-01 tbody tr`
- Ignores rows containing only hyphens (section dividers)
- Writes Hour (HH:00) into column A, then Buy Volume, Sell Volume, Volume, Price
- Expects 23, 24 or 25 rows (one per delivery hour; 23/25 on DST transitions)

Presumed Python 3.12.8
"""

import argparse
import logging
from dataclasses import dataclass
from typing import List, Optional, Sequence

from openpyxl import load_workbook
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeoutError

MARKET_GB = "GB"
PROD_HOUR = 60
EPEX_MARKET = MARKET_GB
EPEX_PRODUCT = PROD_HOUR
CSS_TABLE_SELECTOR = "table.table-01 tbody tr"
EXPECTED_ROW_COUNTS = {23, 24, 25}


@dataclass(frozen=True)
class DARow:
    hour: str                     # "00:00", "01:00", …
    buy_volume: Optional[float]
    sell_volume: Optional[float]
    volume: Optional[float]
    price: Optional[float]


def as_float_or_none(s: str) -> Optional[float]:
    """Convert text to float or return None for '-' / empty."""
    s = (s or "").strip().replace(",", "")
    if not s or s == "-":
        return None
    try:
        return float(s)
    except ValueError:
        return None


def extract_rows_from_dom(page) -> List[DARow]:
    """Extract numeric rows from the rendered EPEX DA table."""
    js = r"""
    (selector) => {
      const rows = [];
      const tableRows = document.querySelectorAll(selector);
      for (const tr of tableRows) {
        const tds = Array.from(tr.querySelectorAll('td')).map(td => (td.textContent || '').trim());
        if (tds.length < 4) continue;
        const meaningful = tds.filter(v => v && v !== '-');
        if (meaningful.length === 0) continue;
        rows.push(tds.slice(0, 4));  // Buy Volume, Sell Volume, Volume, Price
      }
      return rows;
    }
    """
    raw_rows = page.evaluate(js, CSS_TABLE_SELECTOR)
    da_rows: List[DARow] = []
    for i, cells in enumerate(raw_rows):
        nums = [as_float_or_none(x) for x in cells]
        while len(nums) < 4:
            nums.append(None)
        hour_label = f"{i:02d}:00"
        da_rows.append(
            DARow(
                hour=hour_label,
                buy_volume=nums[0],
                sell_volume=nums[1],
                volume=nums[2],
                price=nums[3],
            )
        )
    logging.info("Extracted %d valid rows", len(da_rows))
    return da_rows


def write_rows_to_template(
    template_path: str, out_path: str, rows: Sequence[DARow], start_row: int = 2
) -> None:
    """Write rows into the provided XLSX template."""
    wb = load_workbook(template_path)
    ws = wb.worksheets[0]

    for i, r in enumerate(rows, start=start_row):
        ws.cell(row=i, column=1, value=r.hour)
        ws.cell(row=i, column=2, value=r.buy_volume)
        ws.cell(row=i, column=3, value=r.sell_volume)
        ws.cell(row=i, column=4, value=r.volume)
        ws.cell(row=i, column=5, value=r.price)

    wb.save(out_path)


def run(url: str, template: str, out: str, timeout_ms: int = 30000) -> None:
    logging.info("Navigating to %s", url)
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True, args=["--disable-blink-features=AutomationControlled"])
        context = browser.new_context(user_agent=(
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/120.0.0.0 Safari/537.36"
        ))
        page = context.new_page()

        try:
            page.goto(url, timeout=timeout_ms)
            page.wait_for_selector(CSS_TABLE_SELECTOR, state="attached", timeout=timeout_ms)

            try:
                consent_button = page.query_selector("button:has-text('Accept')") or page.query_selector("button:has-text('Agree')")
                if consent_button:
                    consent_button.click()
                    logging.info("Accepted cookie banner.")
            except Exception:
                pass

            page.wait_for_timeout(1000)
        except PlaywrightTimeoutError as e:
            logging.error("Timed out loading table node: %s", e)
            logging.debug("Page HTML snippet:\n%s", page.content()[:2000])
            raise SystemExit(2)

        rows = extract_rows_from_dom(page)
        if not rows:
            logging.error("No data rows found; verify table selector or structure.")
            raise SystemExit(3)

        if len(rows) not in EXPECTED_ROW_COUNTS:
            logging.warning(
                "Unexpected row count %d (expected 23, 24 or 25); proceeding anyway.",
                len(rows),
            )

        write_rows_to_template(template, out, rows)
        logging.info("Wrote %d rows to %s", len(rows), out)

        context.close()
        browser.close()


def get_epex_url(market: str, date: str, product: int) -> str:
    return (
        f"https://www.epexspot.com/en/market-results"
        f"?market_area={market}&delivery_date={date}"
        f"&modality=Auction&data_mode=table&product={product}"
    )


def parse_args():
    p = argparse.ArgumentParser(description="Scrape EPEX Spot GB Day Ahead Auction table into an XLSX template")
    p.add_argument("--date", required=True, help="The delivery date to fetch (YYYY-MM-DD)")
    p.add_argument("--template", required=True, help="Path to input XLSX template")
    p.add_argument("--out", required=True, help="Output XLSX path")
    p.add_argument("--timeout-ms", type=int, default=30000, help="Page load timeout (ms)")
    p.add_argument("--log-level", default="INFO", help="Logging level")
    return p.parse_args()


if __name__ == "__main__":
    args = parse_args()
    logging.basicConfig(
        level=getattr(logging, args.log_level.upper(), logging.INFO),
        format="%(levelname)s: %(message)s",
    )
    url = get_epex_url(EPEX_MARKET, args.date, EPEX_PRODUCT)
    run(url, args.template, args.out, timeout_ms=args.timeout_ms)
